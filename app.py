from tkinter import *
from tkinter import messagebox
from tkinter import ttk

from PIL import ImageTk, Image
from openpyxl.styles import Border, Side
from openpyxl.workbook import Workbook
from sqlalchemy import text, desc

import db
from models import Producto


class VentanaProducto:
    """"Producto:
        Es una clase para ser usada en GestorProductos
            instanciada en app.py : __name__ == '__main__'
            Maneja la interfaz gráfica para gestionar productos.
        Metodos:
                def db_consulta(self, consulta, parametros=())
                def get_productos(self):
                def validacion_str(self):
                def validacion_num(self):
                def add_producto(self):
                def del_producto(self):
                def edit_producto(self):
                def actualizar_productos(self, nuevo_nombre, antiguo_nombre, nuevo_precio, antiguo_precio):    """

    def __init__(self, root):
        self.ventana = root
        self.ventana.title("App Gestor de Productos")
        self.ventana.resizable(1, 1)
        self.ventana.wm_iconbitmap('recursos/M6_P2_icon.ico')
        self.ventana.configure(background="SteelBlue4")
        # VENTANA PRINCIPAL
        frame = Frame(self.ventana, background="SteelBlue4", bd=2)
        frame.pack()
        # IMAGENES
        imagen_buscar = self.formato_imagen("recursos/buscar.png")
        self.imagen_salir = self.formato_imagen("recursos/salir.png")
        imagen_eliminar = self.formato_imagen("recursos/eliminar.ico")
        imagen_editar = self.formato_imagen("recursos/editar-4.png")
        imagen_guardar = self.formato_imagen("recursos/guardar-2.png")
        self.imagen_excel = self.formato_imagen("recursos/excel.png")
        # MENU
        self.menu = Menu(self.ventana)
        self.ventana.config(menu=self.menu)
        self.boton_menu = Menu(self.menu)
        self.menu.add_cascade(label="Menu", menu=self.boton_menu)
        self.boton_menu.add_command(label="Exportar a Excel", command=self.crear_excel, image=self.imagen_excel,
                                    compound='left')
        self.boton_menu.add_command(label="Salir", command=self.salir, image=self.imagen_salir, compound='left')

        titulo = Label(frame, text="APP GESTOR DE PRODUCTOS",
                       background="SteelBlue4",
                       font=('Calibri', 14, 'bold'))
        titulo.grid(row=0, column=0)

        self.etiqueta_nombre = Label(frame,
                                     text="Nombre: ",
                                     font=('Calibri', 11),
                                     padx=15,
                                     pady=10,
                                     background="SteelBlue4")  # Recibe el LabelFrame
        self.etiqueta_nombre.grid(row=2, column=0)
        self.etiqueta_precio = Label(frame,
                                     text="Precio: ",
                                     font=('Calibri', 11),
                                     padx=15,
                                     pady=10,
                                     background="SteelBlue4")
        self.etiqueta_precio.grid(row=3, column=0)
        self.etiqueta_stock = Label(frame,
                                    text="Stock: ",
                                    font=('Calibri', 11),
                                    padx=15,
                                    pady=10,
                                    background="SteelBlue4")
        self.etiqueta_stock.grid(row=4, column=0)
        self.etiqueta_categoria = Label(frame,
                                        text="Categoria: ",
                                        font=('Calibri', 11),
                                        padx=15,
                                        pady=10,
                                        background="SteelBlue4")
        self.etiqueta_categoria.grid(row=5, column=0)

        self.nombre = Entry(frame,
                            border=4,
                            background="pale turquoise")
        self.nombre.focus()  # foco del raton
        self.nombre.grid(row=2, column=1)
        self.precio = Entry(frame,
                            border=4,
                            background="pale turquoise")
        self.precio.grid(row=3, column=1)
        self.stock = Entry(frame,
                           border=4,
                           background="pale turquoise")
        self.stock.grid(row=4, column=1)
        self.categoria = Entry(frame,
                               border=4,
                               background="pale turquoise")
        self.categoria.grid(row=5, column=1)

        self.boton_aniadir = Button(frame,
                                    text="Guardar Producto",
                                    command=lambda: self.add_producto(
                                        self.nombre,
                                        self.precio,
                                        self.stock,
                                        self.categoria
                                    ),
                                    font=('Arial', 12),
                                    background="medium sea green",
                                    image=imagen_guardar,
                                    compound=LEFT,
                                    borderwidth=3)
        self.boton_aniadir.image = imagen_guardar
        self.boton_aniadir.grid(row=6, columnspan=2, sticky=W + E)

        # Mensaje informativo para el usuario
        self.mensaje = Label(frame,
                             text="",
                             fg="red",
                             bd=0,
                             background="SteelBlue4")
        self.mensaje.grid(row=7, columnspan=3, sticky=W + E)
        self.opcion_precio = IntVar()
        self.boton_check = Checkbutton(frame, text="Ordenar por precio",
                                       variable=self.opcion_precio,
                                       command=self.listar_por_precio,
                                       background="light grey")
        self.boton_check.deselect()
        self.boton_check.grid(row=8, sticky=W + E, padx=5, pady=5)
        # TABLA
        columnas = self.atributos_tabla()
        columnas_a_mostrar = []
        for columna in columnas:
            if columna != 'id':
                columnas_a_mostrar.append(columna)
        print("columnas", columnas)
        print("columnas tipo ", type(columnas))
        frame_tabla = LabelFrame(self.ventana, background="SteelBlue4")
        frame_tabla.pack()

        barra_scroll = Scrollbar(frame_tabla)
        barra_scroll.grid(row=0, column=1, sticky='ns')
        self.tabla = ttk.Treeview(frame_tabla,
                                  height=10,
                                  columns=columnas_a_mostrar,
                                  show='headings',
                                  style="mystyle.Treeview",
                                  yscrollcommand=barra_scroll.set)
        self.tabla.grid(row=0, column=0)

        for columna in columnas_a_mostrar:
            print(" esta es  columna ", columna)
            self.tabla.heading(columna, text=columna, anchor=CENTER)
            self.tabla.column(columna, anchor=CENTER)

        barra_scroll.configure(command=self.tabla.yview)
        # estilos de la tabla
        style = ttk.Style()
        style.configure("mystyle.Treeview",
                        highlightthickness=2,
                        bd=2,
                        font=('Calibri', 11),
                        fieldbackground="SteelBlue4",
                        background="AntiqueWhite3")
        # cabecera de tabla
        style.configure("mystyle.Treeview.Heading", font=('Calibri', 13, 'bold'))
        # tabla sin bordes
        style.layout("mystyle.Treeview", [('mystyle.Treeview.treearea', {'sticky': 'nswe'})])
        # muestra los registros de la bd
        self.get_productos()

        # BOTONES
        botones_frame = LabelFrame(background="SteelBlue4", relief="sunken", bd=0)
        botones_frame.pack()
        # boton eliminar
        self.boton_eliminar = Button(botones_frame, text=' ELIMINAR', command=self.del_producto,
                                     font=('Arial', 12),
                                     background="firebrick3",
                                     image=imagen_eliminar,
                                     compound=LEFT,
                                     borderwidth=3)
        self.boton_eliminar.image = imagen_eliminar
        self.boton_eliminar.grid(row=1, column=1, sticky=W + E, padx=30)

        # boton editar
        self.boton_editar = Button(botones_frame, text=' EDITAR', command=self.edit_producto,
                                   font=('Arial', 12),
                                   background="DarkOrange1",
                                   image=imagen_editar,
                                   compound=LEFT,
                                   borderwidth=3)
        self.boton_editar.image = imagen_editar
        self.boton_editar.grid(row=1, column=0, sticky=W + E, padx=30, pady=15)

        buscar_frame = LabelFrame(background="SteelBlue4", relief="sunken", bd=0)
        buscar_frame.pack()
        self.buscar_nombre = Label(buscar_frame,
                                   text="Nombre: ",
                                   font=('Calibri', 11),
                                   padx=15,
                                   pady=10,
                                   background="SteelBlue4")  # Recibe el LabelFrame
        self.buscar_nombre.grid(row=0, column=0)
        self.buscar_nombre_entry = Entry(buscar_frame,
                                         border=4,
                                         background="pale turquoise")
        self.buscar_nombre_entry.grid(row=0, column=1)
        self.boton_buscar = Button(buscar_frame, text='',
                                   command=lambda: self.buscar_producto(self.buscar_nombre_entry),
                                   font=('Arial', 12),
                                   image=imagen_buscar,
                                   compound=CENTER,
                                   background="yellow",
                                   borderwidth=3)
        self.boton_buscar.image = imagen_buscar
        self.boton_buscar.grid(row=0, column=2)

    def db_consulta(self, consulta):
        """def db_consulta(self, consulta):
            Metodo genérico para consultas a la base de datos productos.db.
            recibe por parametros 'query' y la convierte a text para uso de ORM sqlalchemy. """
        # conexion al la bd
        try:
            resultado = db.session.execute(text(consulta))
            db.session.commit()
            # print("resultado de db_consulta", resultado)
            datos = resultado.fetchall()
            # print("resultado de db_consulta", datos)
            return datos
        except Exception as e:
            print(f"Error en db_consulta: {type(e).__name__}")
            db.session.rollback()

    def formato_imagen(self, ubicacion):
        """def formato_imagen(self, ubicacion):
            Metodo genérico para formato de imagenes
            recibe por parametros 'ubicacion' en directorio de la imagen."""
        try:
            imagen = Image.open(ubicacion)
            imagen_conv = imagen.resize((20, 20))
            return ImageTk.PhotoImage(imagen_conv)
        except EXCEPTION as e:
            print(f"Error en formato_imagen: {type(e).__name__}")

    def atributos_tabla(self):
        """def atributos_tabla(self):
            Método de la clase VentanaProducto.
            Solicita a la tabla producto de productos.bd los nombres de las columas
            y renderiza en: self.tabla = ttk.Treeview. """
        lista_atributos = []
        atributos = db.session.execute(text("PRAGMA table_info(producto)")).fetchall()
        print("atributos: ", atributos)
        for atributo in atributos:
            print("atributos en for: ", atributos)

            lista_atributos.append(atributo[1])

        return lista_atributos

    def get_productos(self):
        """def get_productos(self):
            Método de la clase VentanaProducto.
            Solicita a la tabla producto de productos.bd todos los registros
            y los renderiza en: self.tabla = ttk.Treeview. """
        # Borra lo existete actualmente en la tabla.
        registro_tabla = self.tabla.get_children()
        for fila in registro_tabla:
            self.tabla.delete(fila)
        # Consulta a base de datos
        registros = db.session.query(Producto).order_by(desc("nombre"))
        # print(registros) devuelve un objeto
        # insert de registros en tabla
        for registro in registros:
            print("registros:", registro)
            reg = self.tabla.insert("", 0,
                                    values=(registro.nombre, registro.precio, registro.stock, registro.categoria),
                                    iid=registro.id)

    def listar_por_precio(self):
        """def listar_por_precio(self):
            Método de la clase VentanaProducto.
            Solicita a la tabla producto de productos.bd todos los registros
            y los renderiza en: self.tabla = ttk.Treeview.
            ordenados por precio de forma ASC"""
        self.mensaje['text'] = ''
        registro_tabla = self.tabla.get_children()
        for fila in registro_tabla:
            self.tabla.delete(fila)

        if self.opcion_precio.get() == 1:
            try:
                registros = db.session.query(Producto).order_by("precio")

                for registro in registros:
                    print("registros:", registro)
                    self.tabla.insert("", 0,
                                      values=(registro.nombre, registro.precio, registro.stock, registro.categoria))
            except EXCEPTION as e:
                print(f"Error en del_producto: {type(e).__name__}")

        else:
            self.get_productos()

    def validacion_str(self, entrada):
        """def validacion_str(self):
            Método de la clase VentanaProducto. utilizado para Entry con valores de tipo string
            Recibe de parametro el atributo
            Verifica si el cajon esta vacio T/F
            strip(): elimina espacios vacios del cajon. """
        try:
            return entrada.get().strip() != ""
        except ValueError:
            return False

    def validacion_num(self, num):
        """def validacion_num(self):
            Método de la clase VentanaProducto. utilizado en add_producto() y actualizar_productos()
            Verifica si el campo precio esta vacio T/F. """
        try:
            if float(num.get()):
                return True
            else:
                messagebox.showwarning(title="Atención", message='Ingrese un numero valido')
                return False
        except ValueError:
            return False

    def add_producto(self, nombre, precio, stock, categoria):
        """def add_producto(self):
             Método de la clase VentanaProducto.
             Agrega producto a la tabla producto de productos.db.
             Muestra mensajes informativos.
             Utiliza los metodos:
                validacion_str()
                validacion_num()
                get_productos(): para renderizar nueva tabla. """
        try:
            self.mensaje['text'] = ''
            nombre_valido = self.validacion_str(nombre)
            categoria_valida = self.validacion_str(categoria)
            precio_valido = self.validacion_num(precio) and precio.isdigit()
            stock_valido = self.validacion_num(stock) and stock.isdigit()

            if nombre_valido and precio_valido and stock_valido and categoria_valida:
                nombre_str = str(nombre.get().lower().title())
                precio = round(float(precio.get()), 2)
                stock = round(float(stock.get()), 2)
                categoria_str = str(categoria.get().lower().title())
                producto = Producto(nombre_str, precio, stock, categoria_str)
                db.session.add(producto)
                db.session.commit()
                print("Datos guardados")
                #  Muestra Mensaje informativo
                self.mensaje['text'] = 'Producto {} Guardado'.format(nombre_str)
                messagebox.showinfo(title="Producto Guardado",
                                    message='Producto: {} ¡Añadido con éxito!'.format(nombre_str))
                self.nombre.delete(0, END)  # Borrar el input
                self.precio.delete(0, END)  # Borrar el input
                self.stock.delete(0, END)  # Borrar el input
                self.categoria.delete(0, END)  # Borrar el input
            else:
                no_validos = []
                if not nombre_valido:
                    no_validos.append("NOMBRE")
                if not precio_valido:
                    no_validos.append("PRECIO")
                if not stock_valido:
                    no_validos.append("STOCK")
                if not categoria_valida:
                    no_validos.append("CATEGORIA")
                # print("obligatorios: ", no_validos)
                self.mensaje['text'] = 'Oblgatorio: {}'.format(no_validos)
        except ValueError as e:
            self.mensaje['text'] = f"Error en agregar producto: {type(e).__name__}"
            messagebox.showerror(title="Error", message=f"Error en agregar producto: {type(e).__name__}")
        self.get_productos()

    def del_producto(self):
        """def del_producto(self):
            Método de la clase Producto.
            Elimina registro de la tabla producto de productos.db."""

        self.mensaje['text'] = ''
        seleccion_id = self.tabla.selection()[0]
        print("seleccion: ", seleccion_id)
        if not seleccion_id:
            self.mensaje['text'] = 'Por favor, seleccione un producto'
            raise IndexError
        mensaje_si_no = messagebox.askyesno(title="Eliminar producto",
                                            message="¿Quieres eliminar definitivamente el producto?")

        if mensaje_si_no:
            try:
                producto = db.session.query(Producto).filter(Producto.id == seleccion_id).first()
                self.mensaje['text'] = 'Producto {} eliminado con éxito'.format(producto.nombre)
                db.session.delete(producto)
                db.session.commit()
                self.get_productos()  # Actualizar la tabla de productos

            except IndexError as e:
                print(f"Error en del_producto: {type(e).__name__}")
        else:
            return

    def edit_producto(self):
        """def edit_producto(self):
            Método de la clase VentanaProducto.
            Crea  ventana para modificar un registro de la tabla producto de prductos.db.
            Utiliza metodo:
                    actualizar_productos(): envia 9 parametros"""

        self.mensaje['text'] = ''
        seleccion_id = self.tabla.selection()
        print("seleccion_id", seleccion_id)
        if not seleccion_id:
            self.mensaje['text'] = 'Por favor, seleccione un producto'
            raise IndexError
        try:
            # accedo a una variable de otra ventana
            old_nombre = self.tabla.item(seleccion_id)['values'][0]
            old_precio = self.tabla.item(seleccion_id)['values'][1]
            old_stock = self.tabla.item(seleccion_id)['values'][2]
            old_cateoria = self.tabla.item(seleccion_id)['values'][3]
            # CREAR UNA NUEVA VENTANA
            self.ventana_editar = Toplevel(background="SteelBlue4")
            self.ventana_editar.title("Editar Producto")
            self.ventana_editar.resizable(1, 1)
            self.ventana_editar.wm_iconbitmap('recursos/M6_P2_icon.ico')
            # Label
            titulo = Label(self.ventana_editar,
                           text="Edicion de producto",
                           font=("Calibri", 16, "bold"),
                           background="SteelBlue4")
            titulo.pack()
            # LabelFrame
            frame_ep = LabelFrame(self.ventana_editar,
                                  pady=5,
                                  padx=5,
                                  background="SteelBlue4",
                                  font=('Calibri', 12, 'bold'),
                                  bd=0)
            frame_ep.pack()
            # nombre_antiguo
            self.etiqueta_nombre_antiguo = Label(frame_ep,
                                                 text="Nombre antiguo: ",
                                                 font=('Calibri', 11),
                                                 padx=15,
                                                 pady=10,
                                                 background="SteelBlue4")
            self.etiqueta_nombre_antiguo.grid(row=2, column=0)

            self.input_nombre_antiguo = Entry(frame_ep,
                                              textvariable=StringVar(self.ventana_editar, value=old_nombre),
                                              state='readonly',
                                              border=4,
                                              background="pale turquoise")
            self.input_nombre_antiguo.grid(row=2, column=1)

            # nombre_nuevo
            self.etiqueta_nombre_nuevo = Label(frame_ep,
                                               text="Nombre nuevo: ",
                                               font=('Calibri', 11),
                                               padx=15,
                                               pady=10,
                                               background="SteelBlue4")
            self.etiqueta_nombre_nuevo.grid(row=3, column=0)

            self.input_nombre_nuevo = Entry(frame_ep,
                                            font=("Calibri", 11, "bold"),
                                            border=3,
                                            background="pale green")
            self.input_nombre_nuevo.grid(row=3, column=1)

            # precio_antiguo
            self.etiqueta_precio_antiguo = Label(frame_ep,
                                                 text="Precio antiguo: ",
                                                 font=('Calibri', 11),
                                                 padx=15,
                                                 pady=10,
                                                 background="SteelBlue4")
            self.etiqueta_precio_antiguo.grid(row=4, column=0)

            self.input_precio_antiguo = Entry(frame_ep,
                                              textvariable=StringVar(self.ventana_editar, value=old_precio),
                                              state='readonly',
                                              border=4,
                                              background="pale turquoise")
            self.input_precio_antiguo.grid(row=4, column=1)

            # precio_nuevo
            self.etiqueta_precio_nuevo = Label(frame_ep,
                                               text="Precio nuevo: ",
                                               font=('Calibri', 11),
                                               padx=15,
                                               pady=10,
                                               background="SteelBlue4")
            self.etiqueta_precio_nuevo.grid(row=5, column=0)

            self.input_precio_nuevo = Entry(frame_ep,
                                            font=("Calibri", 11, "bold"),
                                            border=3,
                                            background="pale green")
            self.input_precio_nuevo.grid(row=5, column=1)

            # STOCK_antiguo
            self.etiqueta_stock_antiguo = Label(frame_ep,
                                                text="Stock antiguo: ",
                                                font=('Calibri', 11),
                                                padx=15,
                                                pady=10,
                                                background="SteelBlue4")
            self.etiqueta_stock_antiguo.grid(row=6, column=0)

            self.input_stock_antiguo = Entry(frame_ep,
                                             textvariable=StringVar(self.ventana_editar, value=old_stock),
                                             state='readonly',
                                             border=4,
                                             background="pale turquoise")
            self.input_stock_antiguo.grid(row=6, column=1)

            # stock_nuevo
            self.etiqueta_stock_nuevo = Label(frame_ep,
                                              text="Stock nuevo: ",
                                              font=('Calibri', 11),
                                              padx=15,
                                              pady=10,
                                              background="SteelBlue4")
            self.etiqueta_stock_nuevo.grid(row=7, column=0)

            self.input_stock_nuevo = Entry(frame_ep,
                                           font=("Calibri", 11, "bold"),
                                           border=3,
                                           background="pale green")
            self.input_stock_nuevo.grid(row=7, column=1)
            # Categoria_antigua
            self.etiqueta_categoria_antiguo = Label(frame_ep,
                                                    text="Categoría antigua: ",
                                                    font=('Calibri', 11),
                                                    padx=15,
                                                    pady=10,
                                                    background="SteelBlue4")
            self.etiqueta_categoria_antiguo.grid(row=8, column=0)

            self.input_categoria_antigua = Entry(frame_ep,
                                                 textvariable=StringVar(self.ventana_editar, value=old_cateoria),
                                                 state='readonly',
                                                 border=4,
                                                 background="pale turquoise")
            self.input_categoria_antigua.grid(row=8, column=1)
            # categoria_nueva
            self.etiqueta_categoria_nueva = Label(frame_ep,
                                                  text="Categoría nueva: ",
                                                  font=('Calibri', 11),
                                                  padx=15,
                                                  pady=10,
                                                  background="SteelBlue4")
            self.etiqueta_categoria_nueva.grid(row=9, column=0)

            self.input_categria_nueva = Entry(frame_ep,
                                              font=("Calibri", 11, "bold"),
                                              border=3,
                                              background="pale green")
            self.input_categria_nueva.grid(row=9, column=1)
            # boton actualizar
            self.boton_actualizar = Button(frame_ep,
                                           text="Actualizar Producto",
                                           font=('Arial', 12),
                                           background="medium sea green",
                                           borderwidth=3,
                                           command=lambda: self.actualizar_productos(self.input_nombre_nuevo,
                                                                                     self.input_nombre_antiguo.get(),
                                                                                     self.input_precio_nuevo,
                                                                                     self.input_precio_antiguo.get(),
                                                                                     self.input_stock_nuevo,
                                                                                     self.input_stock_antiguo.get(),
                                                                                     self.input_categria_nueva,
                                                                                     self.input_categoria_antigua.get(),
                                                                                     seleccion_id))
            self.boton_actualizar.grid(row=10, columnspan=2, sticky=W + E)

        except IndexError as e:
            print(f"Error en edit_producto: {type(e).__name__}")
            self.mensaje['text'] = 'Por favor, verifique los nuevos valores'
        except Exception as e:
            print(f"Error en edit_producto: {type(e).__name__}")
            print("Error: vuelva a intentar", e)

    def actualizar_productos(self, nuevo_nombre, antiguo_nombre, nuevo_precio, antiguo_precio, nuevo_stock,
                             antiguo_stock, nueva_categoria, antigua_categoria, seleccion_id):
        """def actualizar_productos(self):
            Método de la clase VentanaProducto.
            Untilizado en la ventana creada en edit_producto()
            Modifica un registro de la tabla producto de prductos.db.
            Recibe 9 parametros:nuevo_nombre, antiguo_nombre, nuevo_precio, antiguo_precio, nuevo_stock,
            antiguo_stock, nueva_categoria, antigua_categoria, seleccion_id) """

        try:
            # Validaciones
            nombre_valido = self.validacion_str(nuevo_nombre)
            categoria_valida = self.validacion_str(nueva_categoria)
            precio_valido = self.validacion_num(nuevo_precio)
            stock_valido = self.validacion_num(nuevo_stock)
            nombre_modificado = nuevo_nombre.get().strip() != "" and nuevo_nombre.get() != antiguo_nombre
            precio_modificado = nuevo_precio.get().strip() != "" and nuevo_precio.get() != antiguo_precio
            stock_modificado = nuevo_stock.get().strip() != "" and nuevo_stock.get() != antiguo_stock
            categoria_modificada = nueva_categoria.get().strip() != "" and nueva_categoria.get() != antigua_categoria
            print(nombre_modificado, precio_modificado, stock_modificado, categoria_modificada)
            # Obtener producto
            seleccion_id_ = seleccion_id[0]
            producto = db.session.query(Producto).filter(Producto.id == seleccion_id_).first()
            # cast
            nombre_str = str(nuevo_nombre.get().strip()).lower().title()
            categoria_str = str(nueva_categoria.get().strip()).lower().title()
            precio_float = round(float(nuevo_precio.get().strip()), 2)
            stock_float =round(float(nuevo_stock.get().strip()), 2)
            if (nombre_modificado and nombre_valido) and (precio_modificado and precio_valido) and (
                    stock_modificado and stock_valido) and (categoria_modificada and categoria_valida):
                producto.nombre = nombre_str
                producto.precio = precio_float
                producto.stock = stock_float
                producto.categoria = categoria_str
                self.mensaje[
                    'text'] = f'El producto {antiguo_nombre} ha sido actualizado con éxito a {nuevo_nombre.get()}'

            ######### MODIFICAR NOMBRE, PRECIO Y CATEGORIA ############
            elif nombre_modificado and precio_modificado and nombre_valido and precio_valido and categoria_modificada and categoria_valida:
                producto.nombre = nombre_str
                producto.precio = precio_float
                producto.categoria = categoria_str
                self.mensaje[
                    'text'] = f'El producto {antiguo_nombre} ha sido actualizado con éxito a {nuevo_nombre.get()}'
                messagebox.showwarning(title=None, message="Has modificado nombre, precio y categoría")

            ######### MODIFICAR NOMBRE Y CATEGORIA ############
            elif nombre_modificado and categoria_modificada and nombre_valido and categoria_valida:
                producto.nombre = nombre_str
                producto.categoria = categoria_str
                self.mensaje[
                    'text'] = f'El producto {antiguo_nombre} ha sido actualizado con éxito a {nuevo_nombre.get()}'
                messagebox.showwarning(title=None, message="Has modificado únicamente el nombre y la categoría")

            ######### MODIFICAR NOMBRE Y PRECIO ############
            elif nombre_modificado and precio_modificado and nombre_valido and precio_valido:
                producto.nombre = nombre_str
                producto.precio = precio_float
                self.mensaje[
                    'text'] = f'El producto {antiguo_nombre} ha sido actualizado con éxito a {nuevo_nombre.get()}'
                messagebox.showwarning(title=None, message="Has modificado únicamente el nombre y el precio")

            ######### MODIFICAR NOMBRE Y STOCK ############
            elif nombre_modificado and stock_modificado and nombre_valido and stock_valido:
                producto.nombre = nombre_str
                producto.stock = stock_float
                self.mensaje[
                    'text'] = f'El producto {antiguo_nombre} ha sido actualizado con éxito a {nuevo_nombre.get()}'
                messagebox.showwarning(title=None, message="Has modificado únicamente el nombre y el stock")

            ######### MODIFICAR PRECIO Y STOCK ############
            elif precio_modificado and stock_modificado and precio_valido and stock_valido:
                producto.precio = precio_float
                producto.stock = stock_float
                self.mensaje[
                    'text'] = f'El producto {antiguo_nombre} ha sido actualizado con éxito'
                messagebox.showwarning(title=None, message="Has modificado únicamente el precio y el stock")

            ######### MODIFICAR PRECIO Y CATEGORIA ############
            elif precio_modificado and precio_valido and categoria_modificada and categoria_valida:
                producto.precio = precio_float
                producto.categoria = categoria_str
                self.mensaje[
                    'text'] = f'El producto {antiguo_nombre} ha sido actualizado con éxito'
                messagebox.showwarning(title=None, message="Has modificado únicamente el precio y la categoría")

            ######### MODIFICAR STOCK Y CATEGORIA ############
            elif stock_modificado and stock_valido and categoria_modificada and categoria_valida:
                producto.stock = stock_float
                producto.categoria = categoria_str
                self.mensaje[
                    'text'] = f'El producto {antiguo_nombre} ha sido actualizado con éxito'
                messagebox.showwarning(title=None, message="Has modificado únicamente el stock y la categoría")

            ######## MODIFICAR UNICAMENTE NOMBRE #########
            elif nombre_modificado and nombre_valido:
                producto.nombre = nombre_str
                self.mensaje[
                    'text'] = f'El producto {antiguo_nombre} ha sido actualizado con éxito a {nuevo_nombre.get()}'
                messagebox.showwarning(title=None, message="Has modificado únicamente el nombre")

            ######## MODIFICAR UNICAMENTE PRECIO #########
            elif precio_modificado and precio_valido:
                producto.precio = precio_float
                self.mensaje['text'] = f'Se ha actualizado el precio de: {antiguo_nombre} precio {nuevo_precio.get()}'
                messagebox.showwarning(title=None, message="Has modificado únicamente el precio")

            ######## MODIFICAR UNICAMENTE STOCK #########
            elif stock_modificado and stock_valido:
                producto.stock = stock_float
                self.mensaje['text'] = f'Se ha actualizado el stock de: {antiguo_nombre} stock: {nuevo_stock.get()}'
                messagebox.showwarning(title=None, message="Has modificado únicamente el stock")

            ######## MODIFICAR UNICAMENTE CATEGORIA #########
            elif categoria_modificada and categoria_valida:
                producto.categoria = categoria_str
                self.mensaje[
                    'text'] = f'Se ha actualizado la categoría de: {antiguo_nombre} categoría: {nueva_categoria.get()}'
                messagebox.showwarning(title=None, message="Has modificado únicamente la categoría")

            else:
                messagebox.showinfo(title=None,
                                    message='No se realizaron cambios en:  {}'.format(antiguo_nombre))

            if nombre_modificado or precio_modificado or stock_modificado or categoria_modificada:
                db.session.commit()
                self.ventana_editar.destroy()
                self.get_productos()

        except IndexError as e:
            print(f"Error en edit_producto: {type(e).__name__}")
            messagebox.showinfo(title=None, message='Verifique los valores ingresados')
            db.session.rollback()

    def salir(self):
        """def salir(self):
            Método de la clase VentanaProducto.
            para cerrar  la app
             """
        mensaje_si_no = messagebox.askyesno(title="Salir", message="¿Quieres salir definitivamente?")
        if mensaje_si_no:
            try:
                root.destroy()
            except Exception as e:
                print(f"Error en del_producto: {type(e).__name__}")

    def buscar_producto(self, buscar_nom):
        """def buscar_producto(self):
            Método de la clase VentanaProducto.
            busca productos en la base de datos segun su nombre.
            Renderiza una nueva tabla con los registros buscados
             """
        try:
            self.mensaje['text'] = ''
            nombre_valido = self.validacion_str(buscar_nom)
            print("nombre_valido", nombre_valido)
            print("buscar_nom", buscar_nom.get())
            if nombre_valido:
                nombre_buscado = buscar_nom.get().strip().lower().title()
                print("nombre_buscado", nombre_buscado)
                productos = db.session.query(Producto).filter(Producto.nombre.ilike(f"{nombre_buscado}%")).all()
                print("buscar productos: ", productos)
                if len(productos) > 0:
                    # CREAR UNA NUEVA VENTANA
                    self.ventana_buscar = Toplevel(background="SteelBlue4")
                    self.ventana_buscar.title("Buscar Producto")
                    self.ventana_buscar.resizable(1, 1)
                    self.ventana_buscar.wm_iconbitmap('recursos/M6_P2_icon.ico')
                    # Label
                    titulo = Label(self.ventana_buscar,
                                   text="Buscar de producto",
                                   font=("Calibri", 16, "bold"),
                                   background="SteelBlue4")
                    titulo.pack()
                    # TABLA
                    columnas = self.atributos_tabla()
                    columnas_a_mostrar = []
                    for columna in columnas:
                        if columna != 'id':
                            columnas_a_mostrar.append(columna)
                    len_colunas = len(columnas)
                    print("columnas", columnas)
                    print("columnas tipo ", type(columnas))
                    frame_tabla_buscar = LabelFrame(self.ventana_buscar, background="SteelBlue4")
                    frame_tabla_buscar.pack()

                    barra_scroll = Scrollbar(frame_tabla_buscar)
                    barra_scroll.grid(row=0, column=1, sticky='ns')
                    self.tabla_buscar = ttk.Treeview(frame_tabla_buscar,
                                                     height=10,
                                                     columns=columnas_a_mostrar,
                                                     show='headings',
                                                     style="mystyle.Treeview",
                                                     yscrollcommand=barra_scroll.set)
                    self.tabla_buscar.grid(row=0, column=0)

                    for columna in columnas_a_mostrar:
                        print(" esta es  columna ", columna)
                        self.tabla_buscar.heading(columna, text=columna, anchor=CENTER)
                        self.tabla_buscar.column(columna, anchor=CENTER)

                    barra_scroll.configure(command=self.tabla_buscar.yview)
                    # estilos de la tabla
                    style = ttk.Style()
                    style.configure("mystyle.Treeview_buscar",
                                    highlightthickness=2,
                                    bd=2,
                                    font=('Calibri', 11),
                                    fieldbackground="SteelBlue4",
                                    background="lavender")
                    # cabecera de tabla
                    style.configure("mystyle.Treeview_buscar.Heading", font=('Calibri', 13, 'bold'))
                    # tabla sin bordes
                    style.layout("mystyle.Treeview_buscar", [('mystyle.Treeview_buscar.treearea', {'sticky': 'nswe'})])
                    # muestra los registros de la bd
                    for registro in productos:
                        print(registro)
                        self.tabla_buscar.insert("", 0, values=(
                            registro.nombre, registro.precio, registro.stock, registro.categoria))
                else:
                    #  Muestra Mensaje informativo
                    messagebox.showinfo(title="Buscar Producto",
                                        message=' No se encontro el producto: {} '.format(nombre_buscado))
                self.buscar_nombre_entry.delete(0, END)  # Borrar el input
        except ValueError as e:
            print(f"Error en edit_producto: {type(e).__name__}")
        return

    def crear_excel(self):
        wb = Workbook()
        hoja_productos = wb.active
        hoja_productos.title = "productos"
        cabecera = self.atributos_tabla()
        registros = db.session.query(Producto).order_by("nombre").all()
        borde = Border(
            left=Side(style='medium', color='000000'),
            right=Side(style='medium', color='000000'),
            top=Side(style='medium', color='000000'),
            bottom=Side(style='medium', color='000000')
        )
        for indice, valor in enumerate(cabecera):
            hoja_productos.cell(row=1, column=indice + 1, value=valor).border = borde
        for row_indice, row in enumerate(registros, start=2):
            hoja_productos.cell(row=row_indice, column=1, value=row.id).border = borde
            hoja_productos.cell(row=row_indice, column=2, value=row.nombre).border = borde
            hoja_productos.cell(row=row_indice, column=3, value=row.precio).border = borde
            hoja_productos.cell(row=row_indice, column=4, value=row.stock).border = borde
            hoja_productos.cell(row=row_indice, column=5, value=row.categoria).border = borde
        wb.save("excel/Gestor_de_productos.xlsx")


if __name__ == '__main__':
    db.Base.metadata.create_all(db.engine)
    root = Tk()
    app = VentanaProducto(root)
    root.mainloop()
    # nada se ejecuta despues del loop
